import openpyxl
import os
import selenium 
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

# Create a webdriver instance
driver = webdriver.Chrome()
driver.maximize_window() 

# Open the Excel file
excel_file_path = #enter the excel path file here; for example: "C:/Excel/Details.xlsx"
#for this line of the code to work make sure your headers in row 1 are as follows:
#username	password	financial year	quarters	period

workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# Find the column indices for "username" and "password"
username_column_index = None
password_column_index = None

for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    for i, cell_value in enumerate(row, start=1):
        if cell_value == "username":
            username_column_index = i
        elif cell_value == "password":
            password_column_index = i

# Check if both columns are found
if username_column_index is None or password_column_index is None:
    print("Column labels 'username' and 'password' not found in the Excel sheet.")
else:
    # Iterate through rows and extract username and password
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username = row[username_column_index - 1]  # Adjust for 0-based indexing
        password = row[password_column_index - 1]  # Adjust for 0-based indexing

        # Navigate to the login page (modify this URL accordingly)
        driver.get('https://services.gst.gov.in/services/login')

        # Find and fill in the username and password fields
        wait = WebDriverWait(driver, 10)  # Adjust the timeout as needed
        username_element = wait.until(EC.visibility_of_element_located((By.ID, "username")))

        if username_element is not None:
            username_element.send_keys(username)
        else:
            print("Username element not found.")
            continue  # Skip this row and move to the next

        password_element = driver.find_element(By.ID, "user_pass")

        if password_element is not None:
            password_element.send_keys(password)
        else:
            print("Password element not found.")
            continue  # Skip this row and move to the next

        time.sleep(6)

        login_button = driver.find_element(By.CSS_SELECTOR, "button[class='btn  btn-primary']")
        login_button.click()

        # Wait for the overlay to disappear

        wait = WebDriverWait(driver, 10)
        return_dashboard = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[class='btn btn-primary pad-l-50 pad-r-50']")))
        time.sleep(5)
        # Click on return dashboard
        return_dashboard = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary pad-l-50 pad-r-50']")
        driver.execute_script("arguments[0].click();", return_dashboard)
        #return_dashboard.click()

        # After clicking the return dashboard button, wait for the page to load
        wait = WebDriverWait(driver, 5)
        wait.until(EC.presence_of_element_located((By.NAME, "fin")))
        
        time.sleep(3)
        
        # Read the financial year value from column C in Excel (cell C1 is labeled as "financial year")
        financial_year_column_index = None
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for i, cell_value in enumerate(row, start=1):
                if cell_value == "financial year":
                    financial_year_column_index = i
                    
        #check if the financial year column is present
        if financial_year_column_index is None:
            print("column label 'financial year' not found in the Excel Sheet")
        else:
            for row in sheet.iter_rows(min_row=2, values_only=True):

        # Extract the financial year from the Excel sheet
                financial_year = row[financial_year_column_index - 1]

        # Find the financial year drop-down element by name
        financial_year_dropdown = driver.find_element(By.NAME, "fin")
        
        time.sleep(3)
        
        wait = WebDriverWait(driver, 6)
        financial_year_dropdown = wait.until(EC.presence_of_element_located((By.NAME, "fin")))

        select = Select(financial_year_dropdown)
        select.select_by_visible_text(financial_year)

        # After selecting the financial year, wait for the quarters drop-down element to become available
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.NAME, "quarter")))

        # Read the quarter value from column D in Excel (cell D1 is labeled as "quarters")
        quarters_column_index = None
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for i, cell_value in enumerate(row, start=1):
                if cell_value == "quarters":
                    quarters_column_index = i
                    
        #check if the financial year column is present
        if quarters_column_index is None:
            print("column label 'Quarter' not found in the Excel Sheet")
        else:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                
        # Extract the quarter from the Excel sheet
                quarters_value = row[quarters_column_index - 1]

        # Find the quarters drop-down element by name
        quarters_dropdown = driver.find_element(By.NAME, "quarter")

        # Select the quarter from the drop-down
        select = Select(quarters_dropdown)
        select.select_by_visible_text(quarters_value)

        # After selecting the quarter, wait for the period (month) drop-down element to become available
        wait = WebDriverWait(driver, 5)
        wait.until(EC.presence_of_element_located((By.NAME, "mon")))

        # Read the period (month) value from column E in Excel (cell E1 is labeled as "period")
        period_column_index = None
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for i, cell_value in enumerate(row, start=1):
                if cell_value == "period":
                    period_column_index = i
                    
        #check if the financial year column is present
        if period_column_index is None:
            print("column label 'Period' not found in the Excel Sheet")
        else:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                
        # Extract the period (month) from the Excel sheet
                    period_value = row[period_column_index - 1]

        # Find the period (month) drop-down element by name
        period_dropdown = driver.find_element(By.NAME, "mon")

        # Select the period (month) from the drop-down
        select = Select(period_dropdown)
        select.select_by_visible_text(period_value)

        # Wait for the overlay to disappear
        wait = WebDriverWait(driver, 10)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))
        
        # Click on search
        search_element = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary srchbtn']")
        driver.execute_script("arguments[0].click();", search_element)
        #search_element.click()
        
         # Wait for the overlay to disappear
    wait = WebDriverWait(driver, 15)
    wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))
    
    time.sleep(10)

    #click on gstr1 button
    gstr_1 = driver.find_element(By.CSS_SELECTOR, "button[data-ng-click='offlinepath(x.return_ty,x.status)']")
    driver.execute_script("arguments[0].click();", gstr_1)
    #gstr_1.click()

    time .sleep(5)
    
# Find the "Generate JSON" button
generate_json = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary']")
driver.execute_script("arguments[0].click();", generate_json)

# You can add a delay if needed between the first and second click
time.sleep(3)  # Adjust the delay as needed

gstr1_dld = driver.find_element(By.XPATH, "//span[contains(text(),'Click here to download - File 1')]")
driver.execute_script("arguments[0].click();", gstr1_dld)

time.sleep(3)

back_element = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-default']")
driver.execute_script("arguments[0].click();", back_element)
#back_element.click()

# ... (previous code)
wait = WebDriverWait(driver, 20)
wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

time.sleep(3)

# Step 1: Load the Subsequent.xlsx file
subsequent_excel_file_path = #enter the excel path file here; for example: "C:/Excel/Subsequent.xlsx"
#for this code to work make sure your header in the excel as follows:
#financial year	  quarters	period

subsequent_workbook = openpyxl.load_workbook(subsequent_excel_file_path)
subsequent_sheet = subsequent_workbook.active

# Find the column indices for "financial year," "quarter," and "period"
financial_year_column_index = 2
quarters_column_index = 3
period_column_index = 4

for row in subsequent_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    for i, cell_value in enumerate(row, start=1):
        if cell_value == "financial year":
            financial_year_column_index = i
        elif cell_value == "quarters":
            quarters_column_index = i
        elif cell_value == "period":
            period_column_index = i

# Check if all three columns are found
if (
    financial_year_column_index is None
    or quarters_column_index is None
    or period_column_index is None
):
    print("Columns 'financial year,' 'quarter,' and 'period' not found in the Subsequent.xlsx sheet.")
else:
    for row in subsequent_sheet.iter_rows(min_row=2, values_only=True):
        financial_year = row[financial_year_column_index - 1]
        quarters_value = row[quarters_column_index - 1]
        period_value = row[period_column_index - 1]
        
        time.sleep(4)
        
        wait = WebDriverWait(driver, 6)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        financial_year_dropdown = driver.find_element(By.NAME, "fin")
        select = Select(financial_year_dropdown)
        select.select_by_visible_text(financial_year)
        
        wait = WebDriverWait(driver, 10)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        quarters_dropdown = driver.find_element(By.NAME, "quarter")
        select = Select(quarters_dropdown)
        select.select_by_visible_text(quarters_value)
        
        wait = WebDriverWait(driver, 100)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        period_dropdown = driver.find_element(By.NAME, "mon")
        select = Select(period_dropdown)
        select.select_by_visible_text(period_value)
        
        wait = WebDriverWait(driver, 15)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        time.sleep(3)

        search_element = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary srchbtn']")
        driver.execute_script("arguments[0].click();", search_element)
        #search_element.click()
        
        time.sleep(5)
 
        gstr_1 = driver.find_element(By.CSS_SELECTOR, "button[data-ng-click='offlinepath(x.return_ty,x.status)']")
        driver.execute_script("arguments[0].click();", gstr_1)
        #gstr_1.click()

        time.sleep(2)
        
        generate_json = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary']")
        driver.execute_script("arguments[0].click();", generate_json)
        #generate_json.click()
          
        time.sleep(3)
        
        gstr1_dld = driver.find_element(By.XPATH, "//span[contains(text(),'Click here to download - File 1')]")
        driver.execute_script("arguments[0].click();", gstr1_dld)

        time.sleep(3)
        
        back_element = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-default']")
        driver.execute_script("arguments[0].click();", back_element)
        #back_element.click()
        
    ########################################### END OF GSTR 1 GENERATION ##########################################################

wait = WebDriverWait(driver, 15)
wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

# Step 1: Load the Subsequent.xlsx file
subsequent_excel_file_path = #enter the excel path file here; for example: "C:/Excel/Subsequent.xlsx"
#for this code to work make sure your header in the excel as follows:
#financial year	  quarters	period

subsequent_workbook = openpyxl.load_workbook(subsequent_excel_file_path)
subsequent_sheet = subsequent_workbook.active

# Find the column indices for "financial year," "quarter," and "period"
financial_year_column_index = 2
quarters_column_index = 3
period_column_index = 4

for row in subsequent_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    for i, cell_value in enumerate(row, start=1):
        if cell_value == "financial year":
            financial_year_column_index = i
        elif cell_value == "quarters":
            quarters_column_index = i
        elif cell_value == "period":
            period_column_index = i

# Check if all three columns are found
if (
    financial_year_column_index is None
    or quarters_column_index is None
    or period_column_index is None
):
    print("Columns 'financial year,' 'quarter,' and 'period' not found in the Subsequent.xlsx sheet.")
else:
    for row in subsequent_sheet.iter_rows(min_row=2, values_only=True):
        financial_year = row[financial_year_column_index - 1]
        quarters_value = row[quarters_column_index - 1]
        period_value = row[period_column_index - 1]
        
        wait = WebDriverWait(driver, 7)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        financial_year_dropdown = driver.find_element(By.NAME, "fin")
        select = Select(financial_year_dropdown)
        select.select_by_visible_text(financial_year)
        
        wait = WebDriverWait(driver, 15)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        quarters_dropdown = driver.find_element(By.NAME, "quarter")
        select = Select(quarters_dropdown)
        select.select_by_visible_text(quarters_value)
        
        wait = WebDriverWait(driver, 15)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        period_dropdown = driver.find_element(By.NAME, "mon")
        select = Select(period_dropdown)
        select.select_by_visible_text(period_value)
        
        wait = WebDriverWait(driver, 15)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        time.sleep(3)

        search_element = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary srchbtn']")
        driver.execute_script("arguments[0].click();", search_element)
        #search_element.click()
        
        time.sleep(5)

        gstr_2a = driver.find_element(By.CSS_SELECTOR, "button[data-ng-click='offlinepath(x.return_ty)']")
        driver.execute_script("arguments[0].click();", gstr_2a)
        
        time.sleep(2)
        
        generate_json = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary']")
        driver.execute_script("arguments[0].click();", generate_json)
        #generate_json.click()
        
        time.sleep(4)  # Add a delay if needed between the first and second click
        
        gstr2a_dld = driver.find_element(By.XPATH, "//span[contains(text(),'Click here to download JSON - File 1')]")
        driver.execute_script("arguments[0],click();", gstr2a_dld)
        
        time.sleep(3)
        
        back_element = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-default']")
        driver.execute_script("arguments[0].click();", back_element)
        #back_element.click()
        
########################################### END OF GSTR 2A GENERATION ##########################################################

wait = WebDriverWait(driver, 15)
wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))
        
subsequent_excel_file_path = #enter the excel path file here; for example: "C:/Excel/Subsequent.xlsx"
#for this code to work make sure your header in the excel as follows:
#financial year	  quarters	period

subsequent_workbook = openpyxl.load_workbook(subsequent_excel_file_path)
subsequent_sheet = subsequent_workbook.active

financial_year_column_index = 2
quarters_column_index = 3
period_column_index = 4

for row in subsequent_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    for i, cell_value in enumerate(row, start=1):
        if cell_value == "financial year":
            financial_year_column_index = i
        elif cell_value == "quarters":
            quarters_column_index = i
        elif cell_value == "period":
            period_column_index = i

# Check if all three columns are found
if (
    financial_year_column_index is None
    or quarters_column_index is None
    or period_column_index is None
):
    print("Columns 'financial year,' 'quarter,' and 'period' not found in the Subsequent.xlsx sheet.")
else:
    for row in subsequent_sheet.iter_rows(min_row=2, values_only=True):
        financial_year = row[financial_year_column_index - 1]
        quarters_value = row[quarters_column_index - 1]
        period_value = row[period_column_index - 1]
        
        wait = WebDriverWait(driver, 15)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        financial_year_dropdown = driver.find_element(By.NAME, "fin")
        select = Select(financial_year_dropdown)
        select.select_by_visible_text(financial_year)
        
        wait = WebDriverWait(driver, 15)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        quarters_dropdown = driver.find_element(By.NAME, "quarter")
        select = Select(quarters_dropdown)
        select.select_by_visible_text(quarters_value)
        
        wait = WebDriverWait(driver, 15)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        period_dropdown = driver.find_element(By.NAME, "mon")
        select = Select(period_dropdown)
        select.select_by_visible_text(period_value)
        
        wait = WebDriverWait(driver, 15)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))

        time.sleep(3)

        search_element = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary srchbtn']")
        driver.execute_script("arguments[0].click();", search_element)
        #search_element.click()
        
        wait = WebDriverWait(driver, 10)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))
 
        #click on gstr2a button
        gstr_2b_xpath = "//div[3]//div[1]//div[1]//div[1]//div[2]//button[1]"
        gstr_2b_button = driver.find_element(By.XPATH, gstr_2b_xpath)
        driver.execute_script("arguments[0].click();", gstr_2b_button)
        #gstr_2b = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary pull-right']")
        #driver.execute_script("arguments[0].click();", gstr_2b)
        
        time.sleep(2)
        
        generate_json = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-primary']")
        driver.execute_script("arguments[0].click();", generate_json)
        #generate_json.click()
        
        time.sleep(3)
        
        back_element = driver.find_element(By.CSS_SELECTOR, "button[class='btn btn-default']")
        driver.execute_script("arguments[0].click();", back_element)
        #back_element.click()
